#!/usr/bin/env python
# coding: utf-8

# In[5]:


import datetime
import csv
import sys
import re
import openpyxl as xl
f_fecha="^[0-9]{2}[/][0-9]{2}[/][0-9]{4}$"
libros={}
obras_autor=[]
obras_genero=[]
obras_año=[]

nombre=input('Escriba el nombre que llevara el archivo:\n').lower()
nombre_csv=(f"{nombre}.csv")
nombre_xl=(f"{nombre}.xlsx")

try:
    with open(nombre_csv,"r", newline="") as archivo:
        lector = csv.reader(archivo)
        next(lector)
        
        for clave, titulo, autor, genero, año, isbn, fecha in lector:
            libros[int(clave)]=(titulo, autor, genero, int(año), int(isbn), fecha)
except FileNotFoundError:
    print("El archivo no se encontró, se procede a trabajar con un conjunto vacío")
    
    while True:
        opcion=int(input(f'Seleccione alguna de las siguientes opciones: \n 1:Subir libro\n 2:Consultas y Reportes\n 3:Salir\n '))
        if opcion==1:
            clave=max(libros, default=0)+1
            while True:
                titulo=input('Ingrese el titulo del libro: ').upper()
                if titulo=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                autor=input('Ingrese el autor del libro: ').upper()                
                if autor=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                genero=input('Ingrese el genero del libro: ').upper()
                if genero=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                try:
                    año=int(input('Ingrese el año en que fue publicado el libro: '))
                except ValueError:
                    print('No son validas las letras en el año ')
                    continue
                else:
                    if año=='':
                        print('No se permite dejar vacios. Intente de nuevo ')
                        continue
                    else:
                        break
            while True:
                try:      
                    isbn=int(input('Ingrese el ISBN del libro: '))
                except ValueError:
                    print('No son validas las letras en el ISBN')
                    continue
                else:
                    if isbn=='':
                        print('No se permite dejar vacios. Intente de nuevo')
                        continue
                    else:
                        break
            while True:
                fecha_adquisicion=input('Ingrese la fecha en que fue adquirido el libro (dd/mm/aaaa): ')
                if fecha_adquisicion=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                if (not bool(re.match(f_fecha, fecha_adquisicion))):
                    print("La fecha inicial no tiene el formato (dd/mm/aaaa)")
                    continue
                else:
                    fecha_adq= datetime.datetime.strptime(fecha_adquisicion, "%d/%m/%Y").date()
                    fecha = (fecha_adq.strftime("%d/%m/%Y"))
                    break
            libros[clave]=[titulo,autor,genero,año,isbn,fecha]
            libros_clave=[(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for clave, datos in libros.items()]
            
        elif opcion==2:
            while True:
                menu=int(input(f'Seleccione alguna de las siguientes opciones:\n1:Consulta de titulo\n2:Reportes\n3:Volver al menu\n'))
                if menu==1:
                    while True:
                        consulta=int(input(f'Seleccione alguna de las siguientes Consultas:\n1:Por titulo\n2:Por ISBN\n3:Salir\n'))
                        if consulta==1:
                            consulta_titulo=input('Ingrese el nombre del titulo a buscar: ').upper()
                            for titulos, autor, genero, año_publicacion, isbn, fecha in libros.values():
                                if titulos==consulta_titulo:
                                    print(f"Datos del Libro\nTitulo={titulos}\nAutor={autor}\nGenero={genero}\nAño de publicacion={año_publicacion}\nISBN={isbn}\nFecha de aquisicion={fecha}")
                        elif consulta==2:
                            consulta_isbn=input('Ingrese el ISBN del libro a consultar')
                            for titulo, autor, genero, año_publicacion, isbns, fecha in libros.values():
                                if isbns==consulta_isbn:
                                    print(f"Datos del Libro\nTitulo={titulo}\nAutor={autor}\nGenero{genero}\nAño de publicacion={año_publicacion}\nISBN={isbns}\nFecha de aquisicion={fecha}")
                        elif consulta==3:
                            break
                        else:
                            print('Opcion no valida')
                            continue

                elif menu==2:
                    while True:
                        reportes=int(input(f"Seleccione por que medio desea realizar los reportes:\n1:Ver catalogo completo\n2:Por autor\n3:Genero\n4:Año publicacion\n5:Volver al menu consultas y reportes\n"))
                        if reportes==1:
                            diccionario = libros.values()
                            print("\n** Catálogo completo ** ")
                            print("Titulo \t\t\tAutor\t\t\tGenero\t\t\tAño de Publicación\t\t\tISBN\t\t\tFecha de Adquisición")
                            print("*" * 120)
                            for titulo, autor, genero, año_publicacion, isbn, fecha in diccionario:
                                print(f"{titulo: <15} | {autor: <15} | {genero: <15} | {año_publicacion: <15} | {isbn:<15} |  {fecha:<15}")
                        
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir '))
                                if exportar==1:
                                    with open(nombre_csv,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                                        grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for clave, datos in libros.items()])
                                    print(f"El archivo fue guardado con el nombre: {nombre_csv}")
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte Completo"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Autor"
                                    hoja["D1"].value = "Genero"
                                    hoja["E1"].value = "Año"
                                    hoja["F1"].value = "ISBN"
                                    hoja["G1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, datos in libros.items():
                                        titulo, autor, genero, año, isbn, fecha = datos
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=autor
                                        hoja.cell(row=renglon, column=4).value=genero
                                        hoja.cell(row=renglon, column=5).value=año
                                        hoja.cell(row=renglon, column=6).value=isbn
                                        hoja.cell(row=renglon, column=7).value=fecha
                                        renglon += 1
                                    nombre_xl.save('Reporte_completo.xlsx')
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break
                        elif reportes==2:
                            print('Los autores disponibles son:')
                            lista_autores=[autor for titulo, autor, genero, año_publicacion, isbn, fecha in libros.values()]
                            set_autores=set(lista_autores)
                            print(set_autores)
                            autor_busqueda=input('Ingrese el nombre del autor: ').upper()
                            print("Titulo \t\t\tAutor\t\t\tGenero\t\t\tAño de Publicación\t\t\tISBN\t\t\tFecha de Adquisición")
                            for clave, titulo, autor, genero, año_publicacion, isbn, fecha in libros_clave:
                                if autor==autor_busqueda:
                                    print(f"{titulo: <15} | {autor: <15} | {genero: <15} | {año_publicacion: <15} | {isbn:<15} |  {fecha:<15}")
                                    obras_autor.append((clave, titulo, autor, genero, año_publicacion, isbn, fecha))
                            autor_aut=(f"Reporte_{autor_busqueda}.csv")
                           
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir '))
                                if exportar==1:
                                    with open(autor_aut,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                                        grabador.writerows([(clave, titulo, autor, genero, año_publicacion, isbn, fecha) for clave, titulo, autor, genero, año_publicacion, isbn, fecha in obras_autor])
                                    print(f"El archivo fue guardado con el nombre: {autor_aut}")
                                  
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    hoja = nombre_xl.active
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte por Autor"                                   
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Autor"
                                    hoja["D1"].value = "Genero"
                                    hoja["E1"].value = "Año"
                                    hoja["F1"].value = "ISBN"
                                    hoja["G1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, titulo, autor, genero, año, isbn, fecha in obras_autor:
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=autor
                                        hoja.cell(row=renglon, column=4).value=genero
                                        hoja.cell(row=renglon, column=5).value=año
                                        hoja.cell(row=renglon, column=6).value=isbn
                                        hoja.cell(row=renglon, column=7).value=fecha
                                        renglon += 1
                                    nombre_xl.save(f"Reporte_{autor_busqueda}.xlsx")
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break

                        elif reportes==3:
                            print('Los generos disponibles son:')
                            lista_generos=[genero for titulo, autor, genero, año_publicacion, isbn, fecha in libros.values()]
                            set_generos=set(lista_generos)
                            print(set_generos)
                            genero_busqueda=input('Ingrese el Genero: ').upper()
                            print("Titulo \t\t\tAutor\t\t\tGenero\t\t\tAño de Publicación\t\t\tISBN\t\t\tFecha de Adquisición")
                            for clave, titulo, autor, genero, año_publicacion, isbn, fecha in libros_clave:
                                if genero==genero_busqueda:
                                    print(f"{titulo: <15} | {autor: <15} | {genero: <15} | {año_publicacion: <15} | {isbn:<15} |  {fecha:<15}")
                                    obras_genero.append((clave, titulo, autor, genero, año_publicacion, isbn, fecha))
                            generos_aut=(f"Reporte_{genero_busqueda}.csv")
                       
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir '))
                                if exportar==1:
                                    with open(generos_aut,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                                        grabador.writerows([(clave, titulo, autor, genero, año_publicacion, isbn, fecha) for clave, titulo, autor, genero, año_publicacion, isbn, fecha in obras_genero])
                                    print(f"El archivo fue guardado con el nombre: {generos_aut}")
                                    
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte por Genero"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Autor"
                                    hoja["D1"].value = "Genero"
                                    hoja["E1"].value = "Año"
                                    hoja["F1"].value = "ISBN"
                                    hoja["G1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, titulo, autor, genero, año, isbn, fecha in obras_genero:
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=autor
                                        hoja.cell(row=renglon, column=4).value=genero
                                        hoja.cell(row=renglon, column=5).value=año
                                        hoja.cell(row=renglon, column=6).value=isbn
                                        hoja.cell(row=renglon, column=7).value=fecha
                                        renglon += 1
                                    nombre_xl.save(f"Reporte_{genero_busqueda}.xlsx")
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break

                        elif reportes==4:
                            print('Los años disponibles son:')
                            lista_años=[año_publicacion for titulo, autor, genero, año_publicacion, isbn, fecha in libros.values()]
                            set_años=set(lista_años)
                            print(set_años)
                            año_busqueda=int(input('Ingrese el año: '))
                            print("Titulo \t\t\tAutor\t\t\tGenero\t\t\tAño de Publicación\t\t\tISBN\t\t\tFecha de Adquisición")
                            for clave, titulo, autor, genero, año_publicacion, isbn, fecha in libros_clave:
                                if int(año_publicacion)==año_busqueda:
                                    print(f"{titulo: <15} | {autor: <15} | {genero: <15} | {año_publicacion: <15} | {isbn:<15} |  {fecha:<15}")
                                    obras_año.append((clave, titulo, autor, genero, año_publicacion, isbn, fecha))
                            años_aut=(f"Reporte_{año_busqueda}.csv")
                           
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir '))
                                if exportar==1:
                                    with open(años_aut,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                                        grabador.writerows([(clave, titulo, autor, genero, año, isbn, fecha) for clave, titulo, autor, genero, año, isbn, fecha in obras_año])
                                    print(f"El archivo fue guardado con el nombre: {años_aut}")
                            
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte por año"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Autor"
                                    hoja["D1"].value = "Genero"
                                    hoja["E1"].value = "Año"
                                    hoja["F1"].value = "ISBN"
                                    hoja["G1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, titulo, autor, genero, año, isbn, fecha in obras_año:
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=autor
                                        hoja.cell(row=renglon, column=4).value=genero
                                        hoja.cell(row=renglon, column=5).value=año
                                        hoja.cell(row=renglon, column=6).value=isbn
                                        hoja.cell(row=renglon, column=7).value=fecha
                                        renglon += 1
                                    nombre_xl.save(f"Reporte_{año_busqueda}.xlsx")
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break
                        elif reportes==5:
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir '))
                                if exportar==1:
                                    with open(nombre_csv,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                                        grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for clave, datos in libros.items()])
                                    print(f"El archivo fue guardado con el nombre: {nombre_csv}")
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte Completo"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Autor"
                                    hoja["D1"].value = "Genero"
                                    hoja["E1"].value = "Año"
                                    hoja["F1"].value = "ISBN"
                                    hoja["G1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, datos in libros.items():
                                        titulo, autor, genero, año_publicacion, isbn, fecha = datos
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=autor
                                        hoja.cell(row=renglon, column=4).value=genero
                                        hoja.cell(row=renglon, column=5).value=año
                                        hoja.cell(row=renglon, column=6).value=isbn
                                        hoja.cell(row=renglon, column=7).value=fecha
                                        renglon += 1
                                    nombre_xl.save('libros.xlsx')
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break
                            break
                        else:
                            print('Opcion no valida')
                            continue
                elif menu==3:
                    break
                else:
                    print('Opcion no valida')
                    break
                    
        elif opcion==3:
            with open(nombre_csv,"w", newline="") as archivo:
                grabador = csv.writer(archivo)
                grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for clave, datos in libros.items()])
            print(f"El archivo fue guardado con el nombre: {nombre_csv}")
            break
        else:
            print('Opcion no valida')
else:
    print('Se encontro el archivo y se procede a usarlo')
    while True:
        opcion=int(input(f'Seleccione alguna de las siguientes opciones: \n 1:Subir libro\n 2:Consultas y Reportes\n 3:Salir\n '))
        if opcion==1:
            clave=max(libros, default=0)+1
            while True:
                titulo=input('Ingrese el titulo del libro: ').upper()
                if titulo=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                autor=input('Ingrese el autor del libro: ').upper()                
                if autor=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                genero=input('Ingrese el genero del libro: ').upper()
                if genero=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                else:
                    break
            while True:
                try:
                    año=int(input('Ingrese el año en que fue publicado el libro: '))
                except ValueError:
                    print('No son validas las letras en el año ')
                    continue
                else:
                    if año=='':
                        print('No se permite dejar vacios. Intente de nuevo ')
                        continue
                    else:
                        break
            while True:
                try:      
                    isbn=int(input('Ingrese el ISBN del libro: '))
                except ValueError:
                    print('No son validas las letras en el ISBN')
                    continue
                else:
                    if isbn=='':
                        print('No se permite dejar vacios. Intente de nuevo')
                        continue
                    else:
                        break
            while True:
                fecha_adquisicion=input('Ingrese la fecha en que fue adquirido el libro (dd/mm/aaaa): ')
                if fecha_adquisicion=='':
                    print('No se permite dejar vacios. Intente de nuevo ')
                    continue
                if (not bool(re.match(f_fecha, fecha_adquisicion))):
                    print("La fecha inicial no tiene el formato (dd/mm/aaaa)")
                    continue
                else:
                    fecha_adq= datetime.datetime.strptime(fecha_adquisicion, "%d/%m/%Y").date()
                    fecha = (fecha_adq.strftime("%d/%m/%Y"))
                    break
            libros[clave]=[titulo,autor,genero,año,isbn,fecha]
            libros_clave=[(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for clave, datos in libros.items()]
            
        elif opcion==2:
            while True:
                menu=int(input(f'Seleccione alguna de las siguientes opciones:\n1:Consulta de titulo\n2:Reportes\n3:Volver al menu\n'))
                if menu==1:
                    while True:
                        consulta=int(input(f'Seleccione alguna de las siguientes Consultas:\n1:Por titulo\n2:Por ISBN\n3:Salir\n'))
                        if consulta==1:
                            consulta_titulo=input('Ingrese el nombre del titulo a buscar: ').upper()
                            for titulos, autor, genero, año_publicacion, isbn, fecha in libros.values():
                                if titulos==consulta_titulo:
                                    print(f"Datos del Libro\nTitulo={titulos}\nAutor={autor}\nGenero={genero}\nAño de publicacion={año_publicacion}\nISBN={isbn}\nFecha de aquisicion={fecha}")
                        elif consulta==2:
                            consulta_isbn=input('Ingrese el ISBN del libro a consultar')
                            for titulo, autor, genero, año_publicacion, isbns, fecha in libros.values():
                                if isbns==consulta_isbn:
                                    print(f"Datos del Libro\nTitulo={titulo}\nAutor={autor}\nGenero{genero}\nAño de publicacion={año_publicacion}\nISBN={isbns}\nFecha de aquisicion={fecha}")
                        elif consulta==3:
                            break
                        else:
                            print('Opcion no valida')
                            continue

                elif menu==2:
                    while True:
                        reportes=int(input(f"Seleccione por que medio desea realizar los reportes:\n1:Ver catalogo completo\n2:Por autor\n3:Genero\n4:Año publicacion\n5:Volver al menu consultas y reportes\n"))
                        if reportes==1:
                            diccionario = libros.values()
                            print("\n** Catálogo completo ** ")
                            print("Titulo \t\t\tAutor\t\t\tGenero\t\t\tAño de Publicación\t\t\tISBN\t\t\tFecha de Adquisición")
                            print("*" * 120)
                            for titulo, autor, genero, año_publicacion, isbn, fecha in diccionario:
                                print(f"{titulo: <15} | {autor: <15} | {genero: <15} | {año_publicacion: <15} | {isbn:<15} |  {fecha:<15}")
                        
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir '))
                                if exportar==1:
                                    with open(nombre_csv,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                                        grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for clave, datos in libros.items()])
                                    print(f"El archivo fue guardado con el nombre: {nombre_csv}")
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte Completo"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Autor"
                                    hoja["D1"].value = "Genero"
                                    hoja["E1"].value = "Año"
                                    hoja["F1"].value = "ISBN"
                                    hoja["G1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, datos in libros.items():
                                        titulo, autor, genero, año, isbn, fecha = datos
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=autor
                                        hoja.cell(row=renglon, column=4).value=genero
                                        hoja.cell(row=renglon, column=5).value=año
                                        hoja.cell(row=renglon, column=6).value=isbn
                                        hoja.cell(row=renglon, column=7).value=fecha
                                        renglon += 1
                                    nombre_xl.save('Reporte_completo.xlsx')
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break
                        elif reportes==2:
                            print('Los autores disponibles son:')
                            lista_autores=[autor for titulo, autor, genero, año_publicacion, isbn, fecha in libros.values()]
                            set_autores=set(lista_autores)
                            print(set_autores)
                            autor_busqueda=input('Ingrese el nombre del autor: ').upper()
                            print("Titulo \t\t\tAutor\t\t\tGenero\t\t\tAño de Publicación\t\t\tISBN\t\t\tFecha de Adquisición")
                            for clave, titulo, autor, genero, año_publicacion, isbn, fecha in libros_clave:
                                if autor==autor_busqueda:
                                    print(f"{titulo: <15} | {autor: <15} | {genero: <15} | {año_publicacion: <15} | {isbn:<15} |  {fecha:<15}")
                                    obras_autor.append((clave, titulo, autor, genero, año_publicacion, isbn, fecha))
                            autor_aut=(f"Reporte_{autor_busqueda}.csv")
                           
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir '))
                                if exportar==1:
                                    with open(autor_aut,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                                        grabador.writerows([(clave, titulo, autor, genero, año_publicacion, isbn, fecha) for clave, titulo, autor, genero, año_publicacion, isbn, fecha in obras_autor])
                                    print(f"El archivo fue guardado con el nombre: {autor_aut}")
                                  
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl.active
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte por Autor"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Autor"
                                    hoja["D1"].value = "Genero"
                                    hoja["E1"].value = "Año"
                                    hoja["F1"].value = "ISBN"
                                    hoja["G1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, titulo, autor, genero, año, isbn, fecha in obras_autor:
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=autor
                                        hoja.cell(row=renglon, column=4).value=genero
                                        hoja.cell(row=renglon, column=5).value=año
                                        hoja.cell(row=renglon, column=6).value=isbn
                                        hoja.cell(row=renglon, column=7).value=fecha
                                        renglon += 1
                                    nombre_xl.save(f"Reporte_{autor_busqueda}.xlsx")
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break

                        elif reportes==3:
                            print('Los generos disponibles son:')
                            lista_generos=[genero for titulo, autor, genero, año_publicacion, isbn, fecha in libros.values()]
                            set_generos=set(lista_generos)
                            print(set_generos)
                            genero_busqueda=input('Ingrese el Genero: ').upper()
                            print("Titulo \t\t\tAutor\t\t\tGenero\t\t\tAño de Publicación\t\t\tISBN\t\t\tFecha de Adquisición")
                            for clave, titulo, autor, genero, año_publicacion, isbn, fecha in libros_clave:
                                if genero==genero_busqueda:
                                    print(f"{titulo: <15} | {autor: <15} | {genero: <15} | {año_publicacion: <15} | {isbn:<15} |  {fecha:<15}")
                                    obras_genero.append((clave, titulo, autor, genero, año_publicacion, isbn, fecha))
                            generos_aut=(f"Reporte_{genero_busqueda}.csv")
                       
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir '))
                                if exportar==1:
                                    with open(generos_aut,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                                        grabador.writerows([(clave, titulo, autor, genero, año_publicacion, isbn, fecha) for clave, titulo, autor, genero, año_publicacion, isbn, fecha in obras_genero])
                                    print(f"El archivo fue guardado con el nombre: {generos_aut}")
                                    
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte por Genero"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Autor"
                                    hoja["D1"].value = "Genero"
                                    hoja["E1"].value = "Año"
                                    hoja["F1"].value = "ISBN"
                                    hoja["G1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, titulo, autor, genero, año, isbn, fecha in obras_genero:
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=autor
                                        hoja.cell(row=renglon, column=4).value=genero
                                        hoja.cell(row=renglon, column=5).value=año
                                        hoja.cell(row=renglon, column=6).value=isbn
                                        hoja.cell(row=renglon, column=7).value=fecha
                                        renglon += 1
                                    nombre_xl.save(f"Reporte_{genero_busqueda}.xlsx")
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break

                        elif reportes==4:
                            print('Los años disponibles son:')
                            lista_años=[año_publicacion for titulo, autor, genero, año_publicacion, isbn, fecha in libros.values()]
                            set_años=set(lista_años)
                            print(set_años)
                            año_busqueda=int(input('Ingrese el año: '))
                            print("Titulo \t\t\tAutor\t\t\tGenero\t\t\tAño de Publicación\t\t\tISBN\t\t\tFecha de Adquisición")
                            for clave, titulo, autor, genero, año_publicacion, isbn, fecha in libros_clave:
                                if int(año_publicacion)==año_busqueda:
                                    print(f"{titulo: <15} | {autor: <15} | {genero: <15} | {año_publicacion: <15} | {isbn:<15} |  {fecha:<15}")
                                    obras_año.append((clave, titulo, autor, genero, año_publicacion, isbn, fecha))
                            años_aut=(f"Reporte_{año_busqueda}.csv")
                           
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir '))
                                if exportar==1:
                                    with open(años_aut,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                                        grabador.writerows([(clave, titulo, autor, genero, año, isbn, fecha) for clave, titulo, autor, genero, año, isbn, fecha in obras_año])
                                    print(f"El archivo fue guardado con el nombre: {años_aut}")
                            
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte por año"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Autor"
                                    hoja["D1"].value = "Genero"
                                    hoja["E1"].value = "Año"
                                    hoja["F1"].value = "ISBN"
                                    hoja["G1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, titulo, autor, genero, año, isbn, fecha in obras_año:
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=autor
                                        hoja.cell(row=renglon, column=4).value=genero
                                        hoja.cell(row=renglon, column=5).value=año
                                        hoja.cell(row=renglon, column=6).value=isbn
                                        hoja.cell(row=renglon, column=7).value=fecha
                                        renglon += 1
                                    nombre_xl.save(f"Reporte_{año_busqueda}.xlsx")
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break
                        elif reportes==5:
                            while True:
                                exportar=int(input('Si quiere exportar seleccione: \n1: CSV \n2: Excel \n3:No exportar y salir '))
                                if exportar==1:
                                    with open(nombre_csv,'w',newline='') as archivo:
                                        grabador = csv.writer(archivo)
                                        grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                                        grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for clave, datos in libros.items()])
                                    print(f"El archivo fue guardado con el nombre: {nombre_csv}")
                                elif exportar==2:
                                    nombre_xl = xl.Workbook()
                                    
                                    hoja = nombre_xl["Sheet"] 
                                    hoja.title = "Reporte Completo"
                                    
                                    hoja["A1"].value = "Clave"
                                    hoja["B1"].value = "Titulo"
                                    hoja["C1"].value = "Autor"
                                    hoja["D1"].value = "Genero"
                                    hoja["E1"].value = "Año"
                                    hoja["F1"].value = "ISBN"
                                    hoja["G1"].value = "Fecha de adquisicion"
                                    
                                    renglon = 2
                                    for clave, datos in libros.items():
                                        titulo, autor, genero, año_publicacion, isbn, fecha = datos
                                        hoja.cell(row=renglon, column=1).value=clave
                                        hoja.cell(row=renglon, column=2).value=titulo
                                        hoja.cell(row=renglon, column=3).value=autor
                                        hoja.cell(row=renglon, column=4).value=genero
                                        hoja.cell(row=renglon, column=5).value=año
                                        hoja.cell(row=renglon, column=6).value=isbn
                                        hoja.cell(row=renglon, column=7).value=fecha
                                        renglon += 1
                                    nombre_xl.save('libros.xlsx')
                                    
                                elif exportar==3:
                                    print("Saliendo")
                                    break
                            break
                        else:
                            print('Opcion no valida')
                            continue
                elif menu==3:
                    break
                else:
                    print('Opcion no valida')
                    break
                    
        elif opcion==3:
            with open(nombre_csv,"w", newline="") as archivo:
                grabador = csv.writer(archivo)
                grabador.writerow(("Clave", "Titulo", "Autor", "Genero", "Año", "ISBN", "Fecha de adquisicion" ))
                grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for clave, datos in libros.items()])
            print(f"El archivo fue guardado con el nombre: {nombre_csv}")
            break
        else:
            print('Opcion no valida')


# In[ ]:




